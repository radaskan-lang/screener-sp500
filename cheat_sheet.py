import pandas as pd
from io import BytesIO
from datetime import datetime


# ─────────────────────────────────────────────
# 📋 CHEAT SHEET — PENSE-BÊTE LUNDI MATIN
# Export Excel avec tous les niveaux clés
# ─────────────────────────────────────────────

def _generate_cheat_sheet(report, market_status):
    """
    Génère un fichier Excel de pense-bête pour le lundi matin.

    Contient pour chaque trade du Top 10 :
    - Ticker, Secteur, Prix actuel
    - Entrée MAX (gap < 1.5%)
    - Stop-loss (marché GTC)
    - Target (vente vendredi)
    - R/R ratio
    - Niveaux Fibonacci (stop et target Fib)
    - Score et convergence
    - Signaux clés
    - Résumé en une ligne
    """
    if report is None or report.empty:
        return None

    try:
        rows = []
        for idx, row in report.iterrows():
            rank    = idx + 1
            ticker  = str(row.get("Ticker", ""))
            sector  = str(row.get("Sector", ""))
            price   = row.get("Prix", None)
            entry   = row.get("Entree", None)
            stop    = row.get("Stop", None)
            target  = row.get("Target", None)
            rr      = row.get("RR_Ratio", None)
            risk_pct= row.get("Risque_Pct", None)
            gain_pct= row.get("Gain_Pct", None)
            conv_n  = row.get("Conv_N", 0)
            score   = row.get("Score_Final", row.get("AI Score Ajuste", 0))
            signal  = str(row.get("AI Signal Ajuste", row.get("AI Signal", "")))

            # Entrée max avec gap 1.5%
            entry_max = round(float(entry) * 1.015, 2) if entry else None

            # Fibonacci
            fib_stop   = row.get("FIB_Stop", None)
            fib_target = row.get("FIB_Target", None)
            fib_rr     = row.get("FIB_RR", None)
            fib_context= str(row.get("FIB_Context", "") or "")
            fib_valid  = row.get("FIB_EntryValid", True)
            fib_warning= str(row.get("FIB_Warning", "") or "")

            # S/R 52 semaines
            sr_high    = row.get("SR_High52w", None)
            sr_low     = row.get("SR_Low52w", None)
            sr_dist    = row.get("SR_DistHigh", None)

            # Relative Strength
            rs_badge   = str(row.get("RS_Badge", "") or "")
            rs_5d      = row.get("RS_Perf5d", None)

            # Signaux principaux
            signals_on = str(row.get("Conv_On", "") or "")
            pattern    = str(row.get("Top_Pattern", "") or "")
            earnings   = str(row.get("Earnings_Badge", "") or "")
            gap_signal = str(row.get("Gap_Signal", "") or "")
            vol_signal = str(row.get("VOL_Signal", "") or "")

            # Résumé une ligne
            resume_parts = []
            if conv_n >= 5: resume_parts.append(f"{conv_n}/6 signaux")
            if rr and float(rr) >= 2: resume_parts.append(f"R/R {rr}:1")
            if pattern and pattern != "—": resume_parts.append(pattern)
            if fib_context == "REBOND_KEY": resume_parts.append("Rebond Fib")
            if fib_context == "RESISTANCE_PROCHE": resume_parts.append("⚠️ Résistance Fib")
            resume = " · ".join(resume_parts)

            rows.append({
                "Rang":            rank,
                "Ticker":          ticker,
                "Secteur":         sector,
                "Signal":          signal,
                "Score":           score,
                "Conv (N/6)":      f"{conv_n}/6",
                "Prix actuel":     price,
                "Entrée MAX":      entry_max,
                "Stop GTC":        stop,
                "Target Vendredi": target,
                "Risque %":        f"-{risk_pct}%" if risk_pct else "",
                "Gain %":          f"+{gain_pct}%" if gain_pct else "",
                "R/R":             f"{rr}:1" if rr else "",
                "Stop Fib":        fib_stop,
                "Target Fib":      fib_target,
                "R/R Fib":         f"{fib_rr}:1" if fib_rr else "",
                "Fib Context":     fib_context,
                "Fib Valide":      "✅ OUI" if fib_valid else "⚠️ NON",
                "Avert. Fib":      fib_warning[:60] if fib_warning and fib_warning != "None" else "",
                "High 52w":        sr_high,
                "Low 52w":         sr_low,
                "Dist. High 52w":  f"{sr_dist}%" if sr_dist else "",
                "RS vs SPY (5j)":  f"{rs_5d}%" if rs_5d else "",
                "RS Badge":        rs_badge,
                "Earnings":        earnings,
                "Gap":             gap_signal[:40] if gap_signal and gap_signal != "None" else "",
                "Volume":          vol_signal[:40] if vol_signal and vol_signal != "None" else "",
                "Pattern":         pattern,
                "Signaux actifs":  signals_on[:80] if signals_on else "",
                "Résumé":          resume,
            })

        df_sheet = pd.DataFrame(rows)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_sheet.to_excel(writer, index=False, sheet_name="Lundi Matin")

            # Mise en forme
            ws = writer.sheets["Lundi Matin"]

            # Largeurs des colonnes
            col_widths = {
                "A": 6,  "B": 8,  "C": 16, "D": 14, "E": 7,
                "F": 10, "G": 12, "H": 12, "I": 10, "J": 16,
                "K": 10, "L": 10, "M": 8,  "N": 10, "O": 12,
                "P": 10, "Q": 16, "R": 12, "S": 40, "T": 10,
                "U": 10, "V": 14, "W": 14, "X": 14, "Y": 16,
                "Z": 40, "AA": 40, "AB": 16, "AC": 80, "AD": 40,
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            # Couleurs de header
            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill = PatternFill(start_color="0A0E1A", end_color="0A0E1A", fill_type="solid")
            header_font = Font(color="00FF88", bold=True, name="Calibri", size=10)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Couleur des lignes
            green_fill  = PatternFill(start_color="001A0F", end_color="001A0F", fill_type="solid")
            yellow_fill = PatternFill(start_color="1A1400", end_color="1A1400", fill_type="solid")
            red_fill    = PatternFill(start_color="1A0000", end_color="1A0000", fill_type="solid")

            for row_idx, row_data in enumerate(rows, start=2):
                score_val = row_data.get("Score", 0) or 0
                fib_ok    = row_data.get("Fib Valide", "✅ OUI")
                if score_val >= 75 and fib_ok == "✅ OUI":
                    fill = green_fill
                elif score_val >= 60:
                    fill = yellow_fill
                else:
                    fill = red_fill
                for col_idx in range(1, len(df_sheet.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        # Ajouter un onglet Info
        with pd.ExcelWriter(output, engine="openpyxl", mode="a") as writer2:
            regime  = market_status.get("regime", "—")
            vix     = market_status.get("vix", "—")
            spy_ma  = market_status.get("spy_vs_ma50", "—")
            info_df = pd.DataFrame([
                {"Info": "Date scan",        "Valeur": datetime.now().strftime("%Y-%m-%d %H:%M")},
                {"Info": "Marché",           "Valeur": regime},
                {"Info": "VIX",              "Valeur": vix},
                {"Info": "SPY vs MA50",      "Valeur": f"{spy_ma}%"},
                {"Info": "Trades sélectionnés", "Valeur": len(rows)},
                {"Info": "Stratégie sortie", "Valeur": "C — Vente vendredi 15h30-15h55 EST"},
                {"Info": "Gap max entrée",   "Valeur": "1.5% max au-dessus du prix vendredi"},
                {"Info": "Stop type",        "Valeur": "Stop au marché GTC — placer immédiatement"},
            ])
            info_df.to_excel(writer2, index=False, sheet_name="Info Marché")

        return output.getvalue()

    except Exception as e:
        return None
